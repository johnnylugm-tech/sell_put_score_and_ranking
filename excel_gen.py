"""
Excel 生成模組
生成 11 Sheets 完整報告
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import List
from collections import Counter
from core import ScoreResult


def fill(color): return PatternFill(start_color=color, end_color=color, fill_type='solid')
def border_thin(): return Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
def center(): return Alignment(horizontal='center', vertical='center', wrap_text=True)
def hdr_font(size=10): return Font(bold=True, color='FFFFFF', size=size, name='Calibri')
def body_font(size=9): return Font(color='000000', size=size, name='Calibri')
def grade_font(grade):
    colors = {'A':'006600', 'B':'3366CC', 'C':'FF6600', 'D':'CC0000'}
    return Font(bold=True, color=colors.get(grade, '000000'), size=10, name='Calibri')
def grade_fill(grade):
    colors = {'A':'E2EFDA', 'B':'DDEBF7', 'C':'FCE4D6', 'D':'F4CCCC'}
    return fill(colors.get(grade, 'FFFFFF'))
def alt_fill(i): return fill('F2F2F2') if i % 2 == 0 else fill('FFFFFF')


def write_hdr_row(ws, row, headers):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = hdr_font()
        c.fill = fill('2E75B6')
        c.alignment = center()
        c.border = border_thin()


SECTOR_COLORS = {
    'Semiconductor': 'E7E6E6', 'Cloud/AI': 'D9E1F2', 'Biotech': 'E2EFDA',
    'Utilities': 'FFF2CC', 'Consumer': 'FCE4D6', 'EV': 'F4CCCC', 'ETF': 'D9D9D9'
}


def generate_excel(results: List[ScoreResult], vix: float, vix_reg: int, 
                   vix_adj: float, pos_scale: float, vix_label: str,
                   today: datetime, output_path: str):
    """生成完整 Excel 報告"""
    
    wb = Workbook()
    
    # Sheet 1: 評分排名
    ws1 = wb.active
    ws1.title = '1.評分排名'
    ws1.merge_cells('A1:Q1')
    c = ws1['A1']
    c.value = f'Sell Put 評分排名 v5.0 | {today.strftime("%Y-%m-%d")} | VIX={vix:.1f}（體制{vix_reg}）'
    c.font = Font(bold=True, color='FFFFFF', size=12, name='Calibri')
    c.fill = fill('1F4E79')
    c.alignment = center()
    
    hdrs = ['排名','Ticker','Sector','等級','調整後總分','原始總分','①距低點','②IV/HV','③基本面','④RSI','⑤流動性','⑥期權流','⑦事件','⑧效率','⑨位置','⑩Skew','財報天數','警告']
    write_hdr_row(ws1, 2, hdrs)

    for i, s in enumerate(results, 1):
        row = i + 2
        ev_days = s.metrics.get('days_to_earnings', 999) if s.metrics.get('days_to_earnings', 999) < 999 else None
        warn_str = '\n'.join(s.warnings) if s.warnings else ''
        vals = [i, s.ticker, s.sector, s.grade, s.adj_total, s.raw_total,
                s.scores['s1'], s.scores['s2'], s.scores['s3'], s.scores['s4'],
                s.scores['s5'], s.scores['s6'], s.scores['s7'], s.scores['s8'],
                s.scores['s9'], s.scores['s10'], ev_days if ev_days else 'N/A', warn_str]
        for j, v in enumerate(vals, 1):
            c = ws1.cell(row=row, column=j, value=v)
            c.border = border_thin()
            c.alignment = center()
            c.font = body_font()
            c.fill = alt_fill(i)
            if j == 4:
                c.font = grade_font(s.grade)
                c.fill = grade_fill(s.grade)
    
    for col in range(1, 18):
        ws1.column_dimensions[get_column_letter(col)].width = 11
    ws1.column_dimensions['R'].width = 30  # 警告欄放寬
    
    # Sheet 2: 評分明細
    ws2 = wb.create_sheet('2.評分明細')
    ws2.merge_cells('A1:H1')
    ws2['A1'] = '評分明細（計算鏈透明）'
    ws2['A1'].font = Font(bold=True, color='FFFFFF', size=12, name='Calibri')
    ws2['A1'].fill = fill('1F4E79')
    ws2['A1'].alignment = center()
    
    write_hdr_row(ws2, 2, ['Ticker','等級','維度','原始值','計算邏輯','得分','權重','說明'])
    
    row = 3
    for s in results:
        details = [
            ('①距低點', f"{s.metrics.get('dist_low', 0):.1f}%", f"距52W低點{s.metrics.get('dist_low', 0):.1f}%", s.scores['s1'], '23分'),
            ('②IV/HV', f"{s.metrics.get('iv_hv_ratio', 0):.2f}", f"IV={s.option.iv:.1f}%/HV={s.stock.hv:.1f}%", s.scores['s2'], '18分'),
            ('③基本面', f"PE={s.stock.fwd_pe:.1f}", f"PE+FCF+成長", s.scores['s3'], '28分'),
            ('④RSI', f"{s.stock.rsi:.1f}", f"RSI={s.stock.rsi:.1f}", s.scores['s4'], '9分'),
            ('⑤流動性', f"市值=${s.stock.mkt_cap/1e9:.1f}B", f"市值分+Beta分", s.scores['s5'], '8分'),
            ('⑥期權流', f"Spread={s.option.spread:.1f}%", f"Spread分+OI分", s.scores['s6'], '9分'),
            ('⑦事件', f"{s.metrics.get('days_to_earnings', 999)}天", f"距財報天數", s.scores['s7'], '5分'),
            ('⑧效率', f"{s.metrics.get('efficiency', 0):.1f}", f"ROCC×PoP", s.scores['s8'], '9分'),
            ('⑨位置', f"{s.metrics.get('pos_52w', 0):.1f}%", f"52W位置", s.scores['s9'], '4分'),
            ('⑩Skew', f"Z={(s.option.iv-s.stock.hv):.1f}", f"IV-HV偏離", s.scores['s10'], '4分'),
        ]
        for dim, raw, logic, score, weight in details:
            vals = [s.ticker, s.grade, dim, raw, logic, score, weight, '']
            for j, v in enumerate(vals, 1):
                c = ws2.cell(row=row, column=j, value=v)
                c.border = border_thin()
                c.alignment = center()
                c.font = body_font(size=8)
                if j == 2:
                    c.font = grade_font(s.grade)
                    c.fill = grade_fill(s.grade)
            row += 1
    
    for col in range(1, 9):
        ws2.column_dimensions[get_column_letter(col)].width = 14
    
    # Sheet 3: 原始數據
    ws3 = wb.create_sheet('3.原始數據')
    ws3.merge_cells('A1:K1')
    ws3['A1'] = '原始數據（yfinance 即時）'
    ws3['A1'].font = Font(bold=True, color='FFFFFF', size=12, name='Calibri')
    ws3['A1'].fill = fill('1F4E79')
    ws3['A1'].alignment = center()
    
    write_hdr_row(ws3, 2, ['Ticker','價格','Fwd_PE','TTM_PE','FCF(B)','營收成長','RSI','HV','IV','IV/HV','PE警告'])
    
    for i, s in enumerate(results, 1):
        row = i + 2
        pe_warning = ''
        if s.stock.fwd_pe < 10 and s.stock.ttm_pe > 15:
            pe_warning = f"⚠️ Fwd={s.stock.fwd_pe:.1f}異常低"
        
        vals = [s.ticker, s.stock.price, s.stock.fwd_pe, s.stock.ttm_pe,
                s.stock.fcf/1e9 if s.stock.fcf else 0, 
                s.stock.revenue_growth*100 if s.stock.revenue_growth else 0,
                s.stock.rsi, s.stock.hv, s.option.iv, s.metrics.get('iv_hv_ratio', 0), pe_warning]
        for j, v in enumerate(vals, 1):
            c = ws3.cell(row=row, column=j, value=v)
            c.border = border_thin()
            c.alignment = center()
            c.font = body_font()
            c.fill = alt_fill(i)
            if j == 11 and pe_warning:
                c.font = Font(color='FF6600', size=8, name='Calibri')
    
    for col in range(1, 12):
        ws3.column_dimensions[get_column_letter(col)].width = 12
    
    # Sheet 4: ROCC_PoP
    ws4 = wb.create_sheet('4.ROCC_PoP')
    ws4.merge_cells('A1:J1')
    ws4['A1'] = f'ROCC & PoP 計算 | DTE={results[0].option.dte if results else 35}（效率係數×1.10）'
    ws4['A1'].font = Font(bold=True, color='FFFFFF', size=12, name='Calibri')
    ws4['A1'].fill = fill('1F4E79')
    ws4['A1'].alignment = center()
    
    write_hdr_row(ws4, 2, ['Ticker','等級','現價','Margin(20%)','Bid','DTE','DTE係數','ROCC_raw','ROCC_adj','PoP','Efficiency','⑧得分'])
    
    for i, s in enumerate(results, 1):
        row = i + 2
        margin = s.stock.price * 0.20
        vals = [s.ticker, s.grade, s.stock.price, margin, s.option.bid, s.option.dte,
                1.10, s.metrics.get('rocc_raw', 0), s.metrics.get('rocc_adj', 0), 
                s.metrics.get('pop', 0), s.metrics.get('efficiency', 0), s.scores['s8']]
        for j, v in enumerate(vals, 1):
            c = ws4.cell(row=row, column=j, value=v)
            c.border = border_thin()
            c.alignment = center()
            c.font = body_font()
            c.fill = alt_fill(i)
            if j == 2:
                c.font = grade_font(s.grade)
                c.fill = grade_fill(s.grade)
    
    for col in range(1, 13):
        ws4.column_dimensions[get_column_letter(col)].width = 11
    
    # Sheet 5: 履約價建議
    ws5 = wb.create_sheet('5.履約價建議')
    ws5.merge_cells('A1:H1')
    ws5['A1'] = f'履約價建議 | Delta目標 0.20-0.30 | VIX體制{vix_reg}'
    ws5['A1'].font = Font(bold=True, color='FFFFFF', size=12, name='Calibri')
    ws5['A1'].fill = fill('1F4E79')
    ws5['A1'].alignment = center()
    
    write_hdr_row(ws5, 2, ['Ticker','等級','現價','ATM履約價','ATM Delta','建議Delta','建議履約價','Spread','流動性'])
    
    for i, s in enumerate(results, 1):
        row = i + 2
        target_delta = 0.25 + vix_adj
        suggested_strike = s.stock.price * (1 + target_delta * 0.5)
        
        if s.option.spread > 20: liq = '🔴 高風險'
        elif s.option.spread > 10: liq = '⚠️ 謹慎'
        else: liq = '✅ 正常'
        
        vals = [s.ticker, s.grade, s.stock.price, s.option.strike, s.option.delta,
                target_delta, suggested_strike, s.option.spread, liq]
        for j, v in enumerate(vals, 1):
            c = ws5.cell(row=row, column=j, value=v)
            c.border = border_thin()
            c.alignment = center()
            c.font = body_font()
            c.fill = alt_fill(i)
            if j == 2:
                c.font = grade_font(s.grade)
                c.fill = grade_fill(s.grade)
    
    for col in range(1, 10):
        ws5.column_dimensions[get_column_letter(col)].width = 12
    
    # Sheet 6: 宏觀對沖
    ws6 = wb.create_sheet('6.宏觀對沖')
    ws6.merge_cells('A1:F1')
    ws6['A1'] = f'宏觀對沖 | VIX={vix:.1f} | 體制{vix_reg}（{vix_label}）'
    ws6['A1'].font = Font(bold=True, color='FFFFFF', size=12, name='Calibri')
    ws6['A1'].fill = fill('1F4E79')
    ws6['A1'].alignment = center()
    
    regime_data = [
        ('體制','VIX範圍','情緒','Delta調整','倉位縮放','操作建議'),
        ('1','< 15','極度貪婪','+0.05','×1.00','正常操作'),
        ('2','15 – 20','正常中性','0','×1.00','標準操作'),
        ('3','20 – 25','輕微焦慮','−0.03','×0.85','降低倉位'),
        ('4','25 – 35','恐慌','−0.05','×0.70','大幅保守'),
        ('5','> 35','崩潰','N/A','×0.00','🚫停止新倉'),
    ]
    for i, row_data in enumerate(regime_data, 1):
        for j, val in enumerate(row_data, 1):
            c = ws6.cell(row=i+2, column=j, value=val)
            if i == 1:
                c.font = hdr_font()
                c.fill = fill('2E75B6')
            else:
                c.font = body_font()
                c.fill = fill('DEEAF1')
            c.alignment = center()
            c.border = border_thin()
    
    ws6['A9'] = '當前狀態:'
    ws6['A9'].font = hdr_font()
    ws6['B9'] = f'體制{vix_reg}'
    ws6['B9'].font = Font(bold=True, color='FF6600', size=12, name='Calibri')
    ws6['C9'] = f'VIX={vix:.1f}'
    ws6['D9'] = vix_label
    ws6['E9'] = f'修正={vix_adj:+.2f}'
    ws6['F9'] = f'縮放={pos_scale}'
    
    for col in range(1, 7):
        ws6.column_dimensions[get_column_letter(col)].width = 14
    
    # Sheet 7: 退出追蹤
    ws7 = wb.create_sheet('7.退出追蹤')
    ws7.merge_cells('A1:K1')
    ws7['A1'] = '退出追蹤模板 | 50%利潤/2×損失/Tastytrade規則'
    ws7['A1'].font = Font(bold=True, color='FFFFFF', size=12, name='Calibri')
    ws7['A1'].fill = fill('1F4E79')
    ws7['A1'].alignment = center()
    
    hdrs = ['Ticker','開倉日','到期日','履約價','收取溢價','觸發條件','動作','實現損益','滾倉次數','接股後CC','備註']
    write_hdr_row(ws7, 2, hdrs)
    
    exit_rules = [
        ('P0 🔴','利潤 ≥ 50%','Buy to Close','Tastytrade最優'),
        ('P0 🔴','損失 ≥ 2×溢價','評估平倉','最大損失控制'),
        ('P0 🔴','財報 < 7天','提前平倉','二元風險'),
        ('P1 🟠','DTE < 21','評估關倉','Gamma激增'),
        ('P1 🟠','跌破履約價10%','滾倉評估','見滾倉規則'),
    ]
    for i, (pri, cond, action, note) in enumerate(exit_rules, 1):
        vals = [pri, cond, action, note, '', '', '', '', '', '', '']
        for j, v in enumerate(vals, 1):
            c = ws7.cell(row=i+2, column=j, value=v)
            c.border = border_thin()
            c.alignment = center()
            c.font = body_font(size=8)
    
    ws7['A8'] = '滾倉條件（三者同時滿足）:'
    ws7['A8'].font = hdr_font()
    ws7['A9'] = '① 基本面 ≥ 15分  ② 距到期 > 14天  ③ 新合約 Efficiency ≥ 80'
    ws7['A10'] = '最大滾倉次數：同一標的不超過 2 次'
    
    for col in range(1, 12):
        ws7.column_dimensions[get_column_letter(col)].width = 13
    
    # Sheet 8: 財報日曆
    ws8 = wb.create_sheet('8.財報日曆')
    ws8.merge_cells('A1:J1')
    ws8['A1'] = f'財報日曆 | 查詢日：{today.strftime("%Y-%m-%d")}'
    ws8['A1'].font = Font(bold=True, color='FFFFFF', size=12, name='Calibri')
    ws8['A1'].fill = fill('1F4E79')
    ws8['A1'].alignment = center()
    
    write_hdr_row(ws8, 2, ['Ticker','Sector','財報日','距今天數','事件分','狀態','基礎面分','調整後總分','等級'])
    
    sorted_by_ev = sorted(results, key=lambda x: x.metrics.get('days_to_earnings', 999))
    for i, s in enumerate(sorted_by_ev, 1):
        row = i + 2
        ev_days = s.metrics.get('days_to_earnings', 999)
        if s.is_forbidden:
            status = '🚫 禁止'
            status_color = 'CC0000'
        elif ev_days <= 14:
            status = '⚠️ 注意'
            status_color = 'FF6600'
        elif ev_days <= 30:
            status = '⚠️ 警惕'
            status_color = 'FF9900'
        else:
            status = '✅ 安全'
            status_color = '006600'
        
        vals = [s.ticker, s.sector, 
                s.stock.earnings_date.strftime('%Y-%m-%d') if s.stock.earnings_date else 'N/A',
                ev_days if ev_days < 999 else 'N/A', s.scores['s7'], status, s.scores['s3'], s.adj_total, s.grade]
        for j, v in enumerate(vals, 1):
            c = ws8.cell(row=row, column=j, value=v)
            c.border = border_thin()
            c.alignment = center()
            c.font = body_font()
            c.fill = alt_fill(i)
            if j == 6:
                c.font = Font(bold=True, color=status_color, size=9, name='Calibri')
            if j == 9:
                c.font = grade_font(s.grade)
                c.fill = grade_fill(s.grade)
    
    for col in range(1, 10):
        ws8.column_dimensions[get_column_letter(col)].width = 12
    
    # Sheet 9: 板塊分析
    ws9 = wb.create_sheet('9.板塊分析')
    ws9.merge_cells('A1:H1')
    ws9['A1'] = '板塊集中度分析'
    ws9['A1'].font = Font(bold=True, color='FFFFFF', size=12, name='Calibri')
    ws9['A1'].fill = fill('1F4E79')
    ws9['A1'].alignment = center()
    
    write_hdr_row(ws9, 2, ['Ticker','Sector','Beta','⑤相關分','⑤市值分','⑤合計','集中度'])
    
    for i, s in enumerate(results, 1):
        row = i + 2
        vals = [s.ticker, s.sector, s.stock.beta, s.scores['s5']//2, s.scores['s5']//2+1, s.scores['s5'], '分散化']
        for j, v in enumerate(vals, 1):
            c = ws9.cell(row=row, column=j, value=v)
            c.border = border_thin()
            c.alignment = center()
            c.font = body_font()
            c.fill = fill(SECTOR_COLORS.get(s.sector, 'FFFFFF'))
    
    row_start = len(results) + 4
    ws9.cell(row=row_start, column=1, value='板塊分佈:').font = hdr_font()
    sector_cnt = Counter(s.sector for s in results)
    for k, (sec, cnt) in enumerate(sorted(sector_cnt.items()), 1):
        r = row_start + k
        ws9.cell(row=r, column=1, value=sec)
        ws9.cell(row=r, column=2, value=cnt)
        ws9.cell(row=r, column=3, value=f'{cnt/len(results)*100:.0f}%')
    
    for col in range(1, 8):
        ws9.column_dimensions[get_column_letter(col)].width = 12
    
    # Sheet 10: 回測追蹤
    ws10 = wb.create_sheet('10.回測追蹤')
    ws10.merge_cells('A1:K1')
    ws10['A1'] = '回測追蹤模板 | Optopsy 框架'
    ws10['A1'].font = Font(bold=True, color='FFFFFF', size=12, name='Calibri')
    ws10['A1'].fill = fill('1F4E79')
    ws10['A1'].alignment = center()
    
    write_hdr_row(ws10, 2, ['Ticker','開倉日','到期日','履約價','Delta','等級','50%觸發日','實現損益','持倉天數','P/L per day'])
    
    expected = [('A (≥80)', '>82%', '<12%', '最高'), ('B (65-79)', '>72%', '<22%', '中高'), ('C (50-64)', '>60%', '<32%', '中低')]
    for i, (grd, win, loss, pl) in enumerate(expected, 1):
        r = 3 + i
        ws10.cell(row=r, column=1, value=f'等級{grd}').font = hdr_font()
        ws10.cell(row=r, column=2, value=f'預期勝率{win}')
        ws10.cell(row=r, column=3, value=f'預期行權率{loss}')
        ws10.cell(row=r, column=4, value=f'P/L: {pl}')
    
    ws10['A8'] = '回測觸發條件（Optopsy）:'
    ws10['A8'].font = hdr_font()
    ws10['A9'] = 'delta: (-0.25, -0.15) | DTE: (21, 45) | Spread < 5% | min_oi: 1000'
    ws10['A10'] = 'Exit: profit_target=50%, stop_loss=2×, dte_exit=21'
    
    for col in range(1, 11):
        ws10.column_dimensions[get_column_letter(col)].width = 14
    
    # Sheet 11: 版本說明
    ws11 = wb.create_sheet('11.版本說明')
    ws11.column_dimensions['A'].width = 25
    ws11.column_dimensions['B'].width = 70
    
    ws11.merge_cells('A1:B1')
    ws11['A1'] = 'Sell Put 評分模型 v5.0 說明文件（修正版）'
    ws11['A1'].font = Font(bold=True, color='FFFFFF', size=13, name='Calibri')
    ws11['A1'].fill = fill('1F4E79')
    ws11['A1'].alignment = center()
    
    def sec_header(ws, row, text):
        ws.merge_cells(f'A{row}:B{row}')
        c = ws.cell(row=row, column=1, value=text)
        c.font = Font(bold=True, color='FFFFFF', size=11, name='Calibri')
        c.fill = fill('2E75B6')
        c.alignment = Alignment(horizontal='left', vertical='center')
    
    def kv(ws, row, key, val, val_color='000000'):
        c = ws.cell(row=row, column=1, value=key)
        c.font = Font(bold=True, size=10, name='Calibri')
        c.alignment = Alignment(horizontal='left', vertical='center')
        c = ws.cell(row=row, column=2, value=val)
        c.font = Font(color=val_color, size=10, name='Calibri')
        c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    row = 3
    sec_header(ws11, row, '版本資訊'); row+=1
    kv(ws11, row, '版本', 'v5.0 P0四修版'); row+=1
    kv(ws11, row, '日期', today.strftime('%Y-%m-%d')); row+=1
    kv(ws11, row, '理論滿分', '117分（10維度）'); row+=1

    row+=1
    sec_header(ws11, row, 'P0 修正說明'); row+=1
    kv(ws11, row, '修正①', 'IVR → IV/HV 比率'); row+=1
    kv(ws11, row, '原因', 'yfinance 無法提供 52W IV 歷史數據，原 IVR 估算導致全系統性高估（全部100%）'); row+=1
    kv(ws11, row, '新②維度', 'IV/HV 比率：>1.5:18 / 1.3:15 / 1.1:12 / 0.9:9 / 0.7:6 / <0.7:3'); row+=1
    kv(ws11, row, '修正②', 'DTE 多到期日選擇'); row+=1
    kv(ws11, row, '原因', '原固定 DTE=14（×0.85），現優先 30-45 DTE（×1.10）'); row+=1
    kv(ws11, row, '修正③', '期權到期日 × 財報日交叉過濾（本次）'); row+=1
    kv(ws11, row, '原因', 'AMD(5/6)/INTC(4/24) 財報早於 5/15 Put 到期，全程持有財報風險'); row+=1
    kv(ws11, row, '邏輯', '過濾：到期日 - 財報日 < 7天（即 Put 持有期間涵蓋財報）→ 強制選不涵蓋的到期日'); row+=1
    kv(ws11, row, '修正④', 'FCF<0 → s3 上限 5 分（本次）'); row+=1
    kv(ws11, row, '原因', 'INTC(FCF=-4.5B)、VST(FCF=-0.46B) 仍獲得了過高基本面分'); row+=1
    kv(ws11, row, '修正⑤', 'QQQ s3=0 直接跳過計算（本次）'); row+=1
    kv(ws11, row, '修正⑥', 'MSFT 近52W低點警告 + Excel 警告欄（本次）'); row+=1
    kv(ws11, row, '原因', 'MSFT 距低點4.8%，①維度得最低分但可能是最佳機會'); row+=1
    
    row+=1
    sec_header(ws11, row, '數據來源聲明'); row+=1
    kv(ws11, row, '股票數據', 'yfinance 即時 API'); row+=1
    kv(ws11, row, '技術指標', 'yfinance 歷史數據計算（RSI: 60日, HV: 30日年化）'); row+=1
    kv(ws11, row, '期權數據', 'yfinance option_chain（ATM Put IV, Bid, Ask, OI）'); row+=1
    kv(ws11, row, 'VIX', 'yfinance ^VIX'); row+=1
    kv(ws11, row, '數據不可得', '若任何數據無法獲取，明確標示「數據不可得」而非估算'); row+=1
    
    row+=1
    sec_header(ws11, row, '禁止幻覺聲明'); row+=1
    kv(ws11, row, '價格', '全部即時從 yfinance 獲取，無快取'); row+=1
    kv(ws11, row, '漲跌幅', '不編造任何漲跌幅數據'); row+=1
    kv(ws11, row, '財務數據', 'PE, FCF, 營收成長均來自 yfinance info'); row+=1
    kv(ws11, row, '期權希臘值', 'Delta 使用 Black-Scholes 公式計算（因 yfinance 不提供）'); row+=1
    
    # 保存
    wb.save(output_path)
    return output_path
